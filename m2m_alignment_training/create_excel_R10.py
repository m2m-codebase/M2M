import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
metric_dir = './metrics'
metrics = []
models = [
    'clip_all_data_no_skip_conn_mse_labse.json',
    'clip_all_data_no_skip_conn_mse_klclip_labse.json',
    'clip_all_data_no_skip_conn_mse_newkl_labse.json',
    'clip_all_data_no_skip_conn_mse_newkl_klclip_labse.json',
]

baseline_models = ['baseline_clip.json']
models = baseline_models + models

for m in models:
    metric_path = os.path.join(metric_dir, m)
    with open(metric_path, 'r') as f:
        metrics.append(json.load(f))

models_data = []
for model, metric in zip(models, metrics):
    model = model.replace('_labse.json', '')
    model = model.replace('.json', '')
    models_data.append({
        "model_name": model,
        "json_data": metric,
    })

# Process JSON data
metrics = []
languages = set()

for model in models_data:
    model_name = model["model_name"]
    json_data = model["json_data"]
    for key, value in json_data.items():
        parts = key.split("/")
        if len(parts) == 4:  # Handle metrics with language information
            metric = parts[1]  # R@1, R@5, R@10
            retrieval_type = parts[2]  # i2t or t2i
            language = parts[3].split("_")[-1]  # Extract language
            languages.add(language)
            metrics.append({
                "Model Name": model_name,
                "Language": language,
                "Retrieval Type": retrieval_type,
                "Metric": metric,
                "Value": value
            })

# Create a DataFrame from metrics
df = pd.DataFrame(metrics)

# Define the desired order for retrieval types and metrics
retrieval_order = ["i2t", "t2i"]
metric_order = ["R@10"]

# Pivot the DataFrame
pivot_df = df.pivot_table(
    index=["Model Name"],
    columns=["Language", "Retrieval Type", "Metric"],
    values="Value"
)

# Reorder the columns to ensure i2t comes before t2i, and metrics follow the specified order
ordered_columns = []
for language in sorted(languages):
    for retrieval_type in retrieval_order:
        for metric in metric_order:
            col = (language, retrieval_type, metric)
            if col in pivot_df.columns:
                ordered_columns.append(col)

pivot_df = pivot_df[ordered_columns]

# Create a workbook
wb = Workbook()
ws = wb.active
ws.title = "Metrics"

# Write merged headers
start_col = 2
languages = sorted(languages)
for language in languages:
    col_start = start_col
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 1)
    ws.cell(row=1, column=col_start, value=language).alignment = Alignment(horizontal="center", vertical="center")
    
    # ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + 2)
    ws.cell(row=2, column=col_start, value="R@10-I2T").alignment = Alignment(horizontal="center", vertical="center")
    
    # ws.merge_cells(start_row=2, start_column=col_start + 3, end_row=2, end_column=col_start + 5)
    ws.cell(row=2, column=col_start + 1, value="R@10-T2I").alignment = Alignment(horizontal="center", vertical="center")
    
    start_col += 2  # Move to the next language group

# Write model data rows
ws.cell(row=2, column=1, value="Model Name")

for row in pivot_df.reset_index().itertuples(index=False):
    ws.append(row)

# Adjust column widths
for i, col in enumerate(ws.columns, start=1):
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(i)].width = max_length + 2


from openpyxl.styles import Font

# Iterate over columns and find the maximum value in each
for col_index, col_cells in enumerate(ws.iter_cols(min_row=4, min_col=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
    # Skip the first column since it contains model names
    column_values = [(cell.row, cell.value) for cell in col_cells if isinstance(cell.value, (int, float))]
    if not column_values:
        continue  # Skip empty columns or non-numeric columns

    # Find the maximum value in the column
    max_value = max(value for _, value in column_values if not np.isnan(value))
    for row, value in column_values:
        if value == max_value:
            # Make the cell bold if it contains the maximum value
            ws.cell(row=row, column=col_index).font = Font(bold=True)


# Save the workbook
file_path = "./excels/recall_metrics_R@10.xlsx"
wb.save(file_path)

